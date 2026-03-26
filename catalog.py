"""
Load product catalog from JOYAS Excel into SQLite.
Maps: Lola = piedras (col G), Barto = taller (col S) + diamantes (col L)
"""
import os
from openpyxl import load_workbook
from dotenv import load_dotenv
import models

load_dotenv()

CATALOG_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    os.getenv("CATALOG_EXCEL", "../Negocio/JOYAS - ANTIQUA 2026 - CORREGIDO.xlsx")
)


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def safe_str(val, default=""):
    return str(val).strip() if val else default


def load_catalog():
    """Read JOYAS Excel and populate products table."""
    abs_path = os.path.abspath(CATALOG_PATH)
    print(f"Loading catalog from: {abs_path}")

    wb = load_workbook(abs_path, read_only=True, data_only=True)
    ws = wb["Joyas"]

    count = 0
    for row in ws.iter_rows(min_row=3, max_col=31, values_only=False):
        vals = [c.value for c in row]
        name = safe_str(vals[1])  # B = Nombre
        if not name:
            continue

        product = {
            "name": name,
            "tipo": safe_str(vals[0]),                    # A = Tipo
            "piedras_desc": safe_str(vals[2]),             # C = Descripcion piedras
            "piedras_total": safe_float(vals[6]),          # G = Total piedras
            "diamantes_desc": safe_str(vals[7]),           # H = Descripcion diamantes
            "diamantes_total": safe_float(vals[11]),       # L = Total diamantes
            "otros_desc": safe_str(vals[12]),              # M = Descripcion otros
            "otros_total": safe_float(vals[14]),           # O = Total otros
            "taller_hechura": safe_float(vals[15]),        # P = Hechura
            "taller_engaste": safe_float(vals[16]),        # Q = Engaste
            "taller_otros": safe_float(vals[17]),          # R = Otros taller
            "taller_total": safe_float(vals[18]),          # S = Total Barto fijo
            "peso_gr": safe_float(vals[19]),               # T = Gr
            "oro_precio_gr": safe_float(vals[20]),         # U = EUR/gr oro
            "oro_total": safe_float(vals[21]),             # V = Total oro
            "cmv": safe_float(vals[22]),                   # W = CMV
            "envio": safe_float(vals[23]),                 # X = Envio
            "pvp": safe_float(vals[27]),                   # AB = PVP
            "iva": safe_float(vals[28]),                   # AC = IVA
            "ingreso": safe_float(vals[29]),               # AD = INGRESO
            "beneficio_bruto": safe_float(vals[30]) if len(vals) > 30 else 0,  # AE
        }
        models.upsert_product(product)
        count += 1
        print(f"  [{count}] {name} ({product['tipo']}) - PVP: {product['pvp']}EUR")

    wb.close()
    print(f"\nCatalog loaded: {count} products")
    return count


def estimate_costs(product_name: str, pvp: float = None, gold_price: float = None):
    """
    Estimate order costs from catalog.
    Lola = piedras (color stones only)
    Barto = taller + diamantes
    """
    product = models.get_product_by_name(product_name)
    if not product:
        return None

    if gold_price is None:
        gold_price = float(os.getenv("GOLD_PRICE_PER_GRAM", "92.0"))

    if pvp is None or pvp == 0:
        pvp = product["pvp"]

    iva = pvp - (pvp / 1.21)
    base = pvp / 1.21
    comision = pvp * 0.021  # Shopify Payments ~2.1%

    # Lola = solo piedras de color
    lola_est = product["piedras_total"]

    # Barto = taller + diamantes
    barto_est = product["taller_total"] + product["diamantes_total"]

    peso_est = product["peso_gr"]
    oro_est = peso_est * gold_price
    cmv_est = lola_est + barto_est + oro_est + product["otros_total"]
    beneficio_est = base - comision - cmv_est

    return {
        "pvp": pvp,
        "iva": round(iva, 2),
        "base_imponible": round(base, 2),
        "comision": round(comision, 2),
        "ingreso_total": round(base - comision, 2),
        "lola_estimado": round(lola_est, 2),
        "barto_estimado": round(barto_est, 2),
        "peso_estimado": round(peso_est, 2),
        "precio_oro_estimado": round(gold_price, 2),
        "oro_total_estimado": round(oro_est, 2),
        "cmv_estimado": round(cmv_est, 2),
        "beneficio_bruto_estimado": round(beneficio_est, 2),
    }


if __name__ == "__main__":
    models.init_db()
    load_catalog()
