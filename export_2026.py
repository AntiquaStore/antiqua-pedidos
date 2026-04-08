"""Export all 2026 Shopify orders to Excel with full details, plus bank transfers not in Shopify."""
import requests, openpyxl, re, os, csv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOKEN = os.getenv("SHOPIFY_ACCESS_TOKEN", "")
STORE = "antiquajoyeria.myshopify.com"
headers = {"X-Shopify-Access-Token": TOKEN}

# Fetch ALL 2026 orders
all_orders = []
url = f"https://{STORE}/admin/api/2025-04/orders.json"
params = {"limit": 250, "status": "any", "created_at_min": "2026-01-01T00:00:00Z"}

while url:
    r = requests.get(url, headers=headers, params=params)
    if r.status_code != 200:
        print(f"Error: {r.status_code}")
        break
    orders = r.json().get("orders", [])
    all_orders.extend(orders)
    link = r.headers.get("Link", "")
    if 'rel="next"' in link:
        for part in link.split(","):
            if 'rel="next"' in part:
                url = part.split("<")[1].split(">")[0]
                params = {}
                break
    else:
        break

print(f"Total orders: {len(all_orders)}")

# Create Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pedidos Antiqua 2026"

hf = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
hfill = PatternFill(start_color="212121", end_color="212121", fill_type="solid")
brd = Border(
    left=Side(style="thin", color="E5E2DC"),
    right=Side(style="thin", color="E5E2DC"),
    top=Side(style="thin", color="E5E2DC"),
    bottom=Side(style="thin", color="E5E2DC"),
)
df = Font(name="Calibri", size=10)
nf = "#,##0.00"

headers_list = [
    "N Pedido", "Fecha", "Cliente", "Email", "Pais",
    "Joya", "Talla", "Variante",
    "PVP (con IVA)", "IVA aplicado", "IVA (%)", "Base imponible",
    "Metodo de pago", "Comision estimada", "Ingreso neto",
    "Estado financiero", "Notas"
]

for i, h in enumerate(headers_list, 1):
    c = ws.cell(row=1, column=i)
    c.value = h
    c.font = hf
    c.fill = hfill
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = brd

row_idx = 2
for order in sorted(all_orders, key=lambda o: o.get("created_at", ""), reverse=True):
    billing = order.get("billing_address") or {}
    customer_name = billing.get("name", "")
    if not customer_name and order.get("customer"):
        cu = order["customer"]
        customer_name = f"{cu.get('first_name','')} {cu.get('last_name','')}".strip()

    email = order.get("email", "")
    country = billing.get("country", "")
    country_code = billing.get("country_code", "")

    fecha = order.get("created_at", "")[:10]
    order_num = order.get("name", "")

    gateways = order.get("payment_gateway_names", [])
    payment_method = ", ".join(gateways) if gateways else "-"

    financial_status = order.get("financial_status", "")
    note = order.get("note", "") or ""

    if country_code in ("ES", ""):
        iva_pct = 21
    elif country_code == "GB":
        iva_pct = 0
    elif country_code in ("PT", "FR", "DE", "IT", "NL", "BE", "AT", "IE", "LU", "GR", "FI", "SE", "DK", "PL", "CZ", "RO", "HU", "BG", "HR", "SK", "SI", "LT", "LV", "EE", "CY", "MT"):
        iva_pct = 21
    else:
        iva_pct = 0

    for item in order.get("line_items", []):
        item_name = item.get("name", "")
        price = float(item.get("price", 0))
        qty = int(item.get("quantity", 1))
        pvp = price * qty

        talla = ""
        variante = ""
        m = re.match(r"^(.+?)\s*-\s*(\d[\d,\.]*)\s*/\s*(.+)$", item_name)
        if m:
            item_name = m.group(1).strip()
            talla = m.group(2).strip()
            variante = m.group(3).strip()
        else:
            m = re.match(r"^(.+?)\s*-\s*(\d[\d,\.]*)$", item_name)
            if m:
                item_name = m.group(1).strip()
                talla = m.group(2).strip()

        if iva_pct > 0:
            base = pvp / (1 + iva_pct / 100)
            iva_amount = pvp - base
        else:
            base = pvp
            iva_amount = 0

        comision = pvp * 0.021 if "shopify" in payment_method.lower() else 0
        ingreso_neto = base - comision

        data = [
            order_num, fecha, customer_name, email,
            f"{country} ({country_code})" if country else country_code,
            item_name, talla, variante,
            pvp, iva_amount, f"{iva_pct}%", base,
            payment_method, comision, ingreso_neto,
            financial_status, note[:100] if note else ""
        ]

        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = val
            cell.font = df
            cell.border = brd
            if col in (9, 10, 12, 14, 15):
                cell.number_format = nf

        row_idx += 1

# ---------------------------------------------------------------------------
# Bank transfers not in Shopify
# ---------------------------------------------------------------------------
parent = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
bank_csv = os.path.join(parent, "Negocio", "Extracto banco 2026.csv")

# Collect all Shopify orders with amounts and customer names for matching
shopify_for_match = []
for o in all_orders:
    total = float(o.get("total_price", 0))
    billing = o.get("billing_address") or {}
    customer = billing.get("name", "")
    if not customer and o.get("customer"):
        cu = o["customer"]
        customer = f"{cu.get('first_name','')} {cu.get('last_name','')}".strip()
    items = [li["name"] for li in o.get("line_items", [])]
    shopify_for_match.append({
        "total": total, "customer": customer, "items": items,
        "order": o.get("name", ""), "used": False
    })

# Parse bank extract
bank_rows = []
with open(bank_csv, "r", encoding="utf-8") as f:
    reader = csv.reader(f, delimiter=";")
    header = next(reader)  # skip header
    for parts in reader:
        if len(parts) < 5:
            continue
        # Skip rows tagged as 2025 in column 5 (index 4)
        if "2025" in str(parts[4]):
            continue
        concepto = parts[0].strip()
        fecha_raw = parts[1].strip()
        importe_raw = parts[2].strip()
        # Only positive amounts (ingresos)
        if not importe_raw.startswith("+"):
            continue
        # Exclude Shopify payouts
        if "TRANSFER. EN DIV" in concepto.upper():
            continue
        # Parse amount: remove +, EUR, dots (thousands sep), comma -> dot
        amt_str = importe_raw.lstrip("+").replace("EUR", "").strip()
        amt_str = amt_str.replace(".", "").replace(",", ".")
        try:
            amount = float(amt_str)
        except ValueError:
            continue
        # Parse date dd/mm/yyyy -> yyyy-mm-dd
        dp = fecha_raw.split("/")
        if len(dp) == 3:
            fecha = f"{dp[2]}-{dp[1].zfill(2)}-{dp[0].zfill(2)}"
        else:
            fecha = fecha_raw
        bank_rows.append({"concepto": concepto, "fecha": fecha, "amount": amount})

# Match against Shopify amounts (tolerance < 1 EUR).
unmatched = []
for br in bank_rows:
    matched = False
    for sm in shopify_for_match:
        if sm["used"]:
            continue
        if abs(br["amount"] - sm["total"]) < 1.0:
            sm["used"] = True
            matched = True
            # Save customer name for matched transfers (useful for reference)
            br["shopify_customer"] = sm["customer"]
            br["shopify_order"] = sm["order"]
            br["shopify_items"] = sm["items"]
            break
    if not matched:
        # Try to find a partial match (customer name in Shopify with similar amount +/- 5%)
        br["shopify_customer"] = ""
        br["shopify_order"] = ""
        br["shopify_items"] = []
        unmatched.append(br)

# Sort unmatched by date
unmatched.sort(key=lambda x: x["fecha"])

print(f"Transferencias bancarias sin match en Shopify: {len(unmatched)}")

if unmatched:
    # Separator row
    sep_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    sep_font = Font(name="Calibri", bold=True, size=11, color="000000")
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers_list))
    sep_cell = ws.cell(row=row_idx, column=1)
    sep_cell.value = "PEDIDOS FUERA DE SHOPIFY (BANCO)"
    sep_cell.font = sep_font
    sep_cell.fill = sep_fill
    sep_cell.alignment = Alignment(horizontal="center")
    for col in range(1, len(headers_list) + 1):
        ws.cell(row=row_idx, column=col).fill = sep_fill
        ws.cell(row=row_idx, column=col).border = brd
    row_idx += 1

    # Write unmatched bank rows
    for br in unmatched:
        amt = br["amount"]
        if amt == 20:
            joya = "Grabado"
            tipo = "Otro"
        elif 20 < amt <= 900:
            joya = "Arreglo u otro"
        else:
            joya = "Pedido fuera de Shopify"

        iva_pct = 21
        base = amt / (1 + iva_pct / 100)
        iva_amount = amt - base
        comision = 0
        ingreso_neto = base - comision

        # Use Shopify customer name if available from partial match, otherwise bank concepto
        cliente = br.get("shopify_customer", "") or br["concepto"]

        data = [
            "BANCO",                    # N Pedido
            br["fecha"],                # Fecha
            cliente,                    # Cliente
            "",                         # Email
            "",                         # Pais
            joya,                       # Joya
            "",                         # Talla
            "",                         # Variante
            amt,                        # PVP (con IVA)
            iva_amount,                 # IVA aplicado
            f"{iva_pct}%",              # IVA (%)
            base,                       # Base imponible
            "Transferencia bancaria",   # Metodo de pago
            comision,                   # Comision estimada
            ingreso_neto,               # Ingreso neto
            "",                         # Estado financiero
            ""                          # Notas
        ]

        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = val
            cell.font = df
            cell.border = brd
            if col in (9, 10, 12, 14, 15):
                cell.number_format = nf

        row_idx += 1

# Column widths
widths = [12, 12, 25, 28, 15, 25, 8, 15, 12, 12, 8, 12, 18, 12, 12, 14, 20]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:Q{row_idx-1}"

# Totals (covers all data rows: Shopify + bank)
ws.cell(row=row_idx, column=8, value="TOTALES").font = Font(name="Calibri", bold=True, size=10)
for col in [9, 10, 12, 14, 15]:
    letter = get_column_letter(col)
    ws.cell(row=row_idx, column=col, value=f"=SUM({letter}2:{letter}{row_idx-1})")
    ws.cell(row=row_idx, column=col).font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=row_idx, column=col).number_format = nf

out = os.path.join(parent, "Negocio", "PEDIDOS-ANTIQUA-2026-COMPLETO.xlsx")
wb.save(out)
print(f"Guardado: {out}")
print(f"Filas de datos: {row_idx - 2}")
