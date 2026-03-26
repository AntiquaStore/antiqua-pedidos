"""
Bidirectional sync between SQLite and pedidos 2026.xlsx (Facturas sheet).
Column mapping A-Z matches the Excel exactly.
"""
import os, datetime
from openpyxl import load_workbook
from dotenv import load_dotenv
import models

load_dotenv()

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    os.getenv("PEDIDOS_EXCEL", "../Negocio/pedidos 2026.xlsx")
)

# Column mapping: Excel col letter -> DB field
COL_MAP = {
    "A": "n_joya",
    "B": "product_name",        # Joya vendida
    "C": "fecha_pedido",
    "D": "pvp",
    "E": "iva",
    "F": "base_imponible",
    "G": "fecha_ingreso",
    "H": "comision",
    "I": "ingreso_total",
    "J": "lola_estimado",
    "K": "lola_real",
    "L": "fecha_cobro_lola",
    "M": "barto_estimado",
    "N": "barto_real",
    "O": "fecha_cobro_barto",
    "P": "peso_estimado",
    "Q": "peso_real",
    "R": "precio_oro_estimado",
    "S": "precio_oro_real",
    "T": "oro_total_estimado",
    "U": "oro_total_real",
    "V": "cmv_estimado",
    "W": "cmv_real",
    "X": "envio_packaging",
    "Y": "beneficio_bruto_estimado",
    "Z": "beneficio_bruto_real",
}


def read_orders_from_excel():
    """Read existing orders from Excel into a list of dicts."""
    path = os.path.abspath(EXCEL_PATH)
    if not os.path.exists(path):
        print(f"Excel not found: {path}")
        return []

    wb = load_workbook(path, data_only=True)
    ws = wb["Facturas"]
    orders = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=26, values_only=True), start=2):
        if not row[1]:  # B = Joya vendida (skip empty rows)
            continue
        data = {}
        for col_idx, (letter, field) in enumerate(COL_MAP.items()):
            val = row[col_idx]
            if val is not None:
                if "fecha" in field:
                    data[field] = str(val)[:10] if val else ""
                elif isinstance(val, (int, float)):
                    data[field] = float(val)
                else:
                    data[field] = str(val)
            else:
                data[field] = None
        data["_excel_row"] = row_idx
        orders.append(data)

    wb.close()
    return orders


def import_excel_to_db():
    """Import orders from Excel into SQLite (initial load)."""
    orders = read_orders_from_excel()
    count = 0
    for i, order_data in enumerate(orders):
        row_num = order_data.pop("_excel_row", 0)
        # Use row number as a unique ID since Excel orders don't have Shopify IDs
        if not order_data.get("product_name"):
            continue

        db_data = {
            "shopify_order_id": f"excel-{row_num}",
            "product_name": order_data.get("product_name", ""),
            "status": "completado" if order_data.get("beneficio_bruto_real") else "nuevo",
        }
        # Merge non-None values
        for k, v in order_data.items():
            if v is not None and k not in ("_excel_row",):
                db_data[k] = v

        models.upsert_order(db_data)
        count += 1

    print(f"Imported {count} orders from Excel")
    return count


def export_order_to_excel(order: dict):
    """Write a single order back to Excel, appending or updating."""
    path = os.path.abspath(EXCEL_PATH)
    if not os.path.exists(path):
        print(f"Excel not found: {path}")
        return False

    wb = load_workbook(path)
    ws = wb["Facturas"]

    # Find existing row by product name + date, or append
    target_row = None
    for row_idx in range(2, ws.max_row + 1):
        cell_b = ws.cell(row=row_idx, column=2).value
        cell_c = ws.cell(row=row_idx, column=3).value
        if cell_b and str(cell_b).strip().lower() == str(order.get("product_name", "")).strip().lower():
            if cell_c:
                excel_date = str(cell_c)[:10]
                order_date = str(order.get("fecha_pedido", ""))[:10]
                if excel_date == order_date:
                    target_row = row_idx
                    break

    if not target_row:
        target_row = ws.max_row + 1

    # Write values
    col_to_field = list(COL_MAP.values())
    for col_idx, field in enumerate(col_to_field, start=1):
        val = order.get(field)
        if val is not None:
            ws.cell(row=target_row, column=col_idx, value=val)

    wb.save(path)
    wb.close()
    return True


def export_all_to_excel():
    """Export all DB orders to Excel."""
    orders = models.get_all_orders()
    count = 0
    for order in orders:
        if export_order_to_excel(order):
            count += 1
    return count


if __name__ == "__main__":
    models.init_db()
    import_excel_to_db()
