"""Export all 2026 orders to Excel - V2: correct, no duplicates, real commissions."""
import requests, openpyxl, re, os, csv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOKEN = os.getenv("SHOPIFY_ACCESS_TOKEN", "")
STORE = "antiquajoyeria.myshopify.com"

# If no env var, read from .env
if not TOKEN:
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                if line.startswith("SHOPIFY_ACCESS_TOKEN"):
                    TOKEN = line.split("=", 1)[1].strip().strip('"').strip("'")

headers_api = {"X-Shopify-Access-Token": TOKEN}

# ============================================================
# 1. FETCH ALL SHOPIFY ORDERS 2026
# ============================================================
all_orders = []
url = f"https://{STORE}/admin/api/2025-04/orders.json"
params = {"limit": 250, "status": "any", "created_at_min": "2026-01-01T00:00:00Z"}
while url:
    r = requests.get(url, headers=headers_api, params=params)
    if r.status_code != 200:
        print(f"Error: {r.status_code}")
        break
    all_orders.extend(r.json().get("orders", []))
    link = r.headers.get("Link", "")
    if 'rel="next"' in link:
        for part in link.split(","):
            if 'rel="next"' in part:
                url = part.split("<")[1].split(">")[0]
                params = {}
                break
    else:
        break

print(f"Shopify orders: {len(all_orders)}")

# ============================================================
# 2. PARSE BANK CSV - only transfers NOT in Shopify
# ============================================================
parent = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
bank_csv = os.path.join(parent, "Negocio", "Extracto banco 2026.csv")

# Build set of Shopify order amounts for matching
shopify_amounts = []
for o in all_orders:
    total = float(o.get("total_price", 0))
    shopify_amounts.append(total)

# Parse bank
bank_extra = []  # transfers not in Shopify
with open(bank_csv, "r", encoding="utf-8") as f:
    reader = csv.reader(f, delimiter=";")
    for row in reader:
        if len(row) < 4:
            continue
        concepto = row[0].strip()
        fecha_raw = row[1].strip()
        importe_raw = row[2].strip()
        extra = row[4].strip() if len(row) > 4 else ""

        # Skip 2025
        if "2025" in extra:
            continue

        # Only positive (ingresos)
        amt_str = importe_raw.replace("EUR", "").replace(".", "").replace(",", ".").strip()
        try:
            amount = float(amt_str)
        except:
            continue
        if amount <= 0:
            continue

        # Skip Shopify payouts (TRANSFER. EN DIV.)
        if "TRANSFER. EN DIV" in concepto:
            continue

        # Parse date
        dp = fecha_raw.split("/")
        if len(dp) == 3:
            fecha = f"{dp[2]}-{dp[1].zfill(2)}-{dp[0].zfill(2)}"
        else:
            fecha = fecha_raw

        # Try to match with Shopify by amount
        matched = False
        for i, sa in enumerate(shopify_amounts):
            if abs(amount - sa) < 1.0:
                shopify_amounts.pop(i)
                matched = True
                break

        if not matched:
            # Classify
            if amount == 20:
                tipo = "Grabado"
            elif amount < 600:
                tipo = "Arreglo u otro"
            else:
                tipo = "Pedido fuera de Shopify"

            bank_extra.append({
                "fecha": fecha,
                "concepto": concepto,
                "amount": amount,
                "tipo": tipo,
            })

print(f"Bank transfers not in Shopify: {len(bank_extra)}")

# ============================================================
# 3. CREATE EXCEL
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pedidos Antiqua 2026"

hf = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
hfill = PatternFill(start_color="212121", end_color="212121", fill_type="solid")
brd = Border(
    left=Side(style="thin", color="E5E2DC"),
    right=Side(style="thin", color="E5E2DC"),
    top=Side(style="thin", color="E5E2DC"),
    bottom=Side(style="thin", color="E5E2DC"),
)
df = Font(name="Calibri", size=10)
bf = Font(name="Calibri", bold=True, size=10)
nf = "#,##0.00"

cols = [
    "N Pedido", "Fecha", "Cliente", "Email", "Pais",
    "Joya", "Talla", "Variante",
    "PVP (con IVA)", "IVA (%)", "IVA importe", "Base imponible",
    "Metodo de pago", "Comision real", "Ingreso neto",
    "Estado", "Notas"
]

for i, h in enumerate(cols, 1):
    c = ws.cell(row=1, column=i)
    c.value = h
    c.font = hf
    c.fill = hfill
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = brd

# ============================================================
# 4. WRITE SHOPIFY ORDERS (1 row per line item, NO duplicates)
# ============================================================
row_idx = 2
for order in sorted(all_orders, key=lambda o: o.get("created_at", ""), reverse=True):
    billing = order.get("billing_address") or {}
    customer_name = billing.get("name", "")
    if not customer_name and order.get("customer"):
        cu = order["customer"]
        customer_name = f"{cu.get('first_name', '')} {cu.get('last_name', '')}".strip()

    email = order.get("email", "")
    country = billing.get("country", "")
    country_code = billing.get("country_code", "")
    fecha = order.get("created_at", "")[:10]
    order_num = order.get("name", "")
    financial_status = order.get("financial_status", "")
    note = order.get("note", "") or ""

    # Payment method - single string, no duplicates
    gateways = order.get("payment_gateway_names", [])
    gateway_str = ", ".join(dict.fromkeys(gateways)) if gateways else "-"

    # Determine IVA rate
    if country_code in ("ES", ""):
        iva_pct = 21
    elif country_code == "GB":
        iva_pct = 0
    elif country_code in ("PT", "FR", "DE", "IT", "NL", "BE", "AT", "IE", "LU",
                          "GR", "FI", "SE", "DK", "PL", "CZ", "RO", "HU", "BG",
                          "HR", "SK", "SI", "LT", "LV", "EE", "CY", "MT"):
        iva_pct = 21
    else:
        iva_pct = 0

    # Real commission: Shopify Payments Spain = 1.4% + 0.25 EUR per transaction
    # PayPal = ~3.4% + 0.35 EUR
    # Transfer / Cash = 0
    order_total = float(order.get("total_price", 0))
    is_shopify_payments = "shopify_payments" in [g.lower() for g in gateways]
    is_paypal = "paypal" in [g.lower() for g in gateways]

    if is_shopify_payments:
        comision_order = order_total * 0.014 + 0.25
    elif is_paypal:
        comision_order = order_total * 0.034 + 0.35
    else:
        comision_order = 0

    # Distribute commission proportionally across line items
    line_items = order.get("line_items", [])
    total_items_value = sum(float(li.get("price", 0)) * int(li.get("quantity", 1)) for li in line_items)

    for item in line_items:
        item_name = item.get("name", "")
        price = float(item.get("price", 0))
        qty = int(item.get("quantity", 1))
        pvp = price * qty

        # Parse talla/variant
        talla = ""
        variante = ""
        m = re.match(r"^(.+?)\s*-\s*(\d[\d,\.]*)\s*/\s*(.+)$", item_name)
        if m:
            item_name = m.group(1).strip()
            talla = m.group(2).strip()
            variante = m.group(3).strip()
        else:
            m = re.match(r"^(.+?)\s*-\s*(\d[\d,\.]*)$", item_name)
            if m:
                item_name = m.group(1).strip()
                talla = m.group(2).strip()

        # IVA
        if iva_pct > 0:
            base = pvp / (1 + iva_pct / 100)
            iva_amount = pvp - base
        else:
            base = pvp
            iva_amount = 0

        # Proportional commission
        if total_items_value > 0:
            comision = comision_order * (pvp / total_items_value)
        else:
            comision = 0

        ingreso_neto = base - comision

        data = [
            order_num, fecha, customer_name, email,
            f"{country} ({country_code})" if country else country_code,
            item_name, talla, variante,
            pvp, f"{iva_pct}%", iva_amount, base,
            gateway_str, round(comision, 2), round(ingreso_neto, 2),
            financial_status, note[:100] if note else ""
        ]

        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = val
            cell.font = df
            cell.border = brd
            if col in (9, 11, 12, 14, 15):
                cell.number_format = nf

        row_idx += 1

shopify_end_row = row_idx - 1

# ============================================================
# 5. SEPARATOR + BANK-ONLY ENTRIES
# ============================================================
if bank_extra:
    sep_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(cols))
    sep_cell = ws.cell(row=row_idx, column=1)
    sep_cell.value = "INGRESOS FUERA DE SHOPIFY (BANCO)"
    sep_cell.font = Font(name="Calibri", bold=True, size=11)
    sep_cell.fill = sep_fill
    sep_cell.alignment = Alignment(horizontal="center")
    for col in range(1, len(cols) + 1):
        ws.cell(row=row_idx, column=col).fill = sep_fill
        ws.cell(row=row_idx, column=col).border = brd
    row_idx += 1

    for br in sorted(bank_extra, key=lambda x: x["fecha"], reverse=True):
        amt = br["amount"]
        iva_pct = 21
        base = amt / (1 + iva_pct / 100)
        iva_amount = amt - base

        data = [
            "BANCO", br["fecha"], br["concepto"], "", "",
            br["tipo"], "", "",
            amt, f"{iva_pct}%", iva_amount, base,
            "Transferencia bancaria", 0, base,
            "", ""
        ]

        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = val
            cell.font = df
            cell.border = brd
            if col in (9, 11, 12, 14, 15):
                cell.number_format = nf

        row_idx += 1

# ============================================================
# 6. TOTALS
# ============================================================
totals_row = row_idx
ws.cell(row=totals_row, column=8, value="TOTALES SHOPIFY").font = bf
for col in [9, 11, 12, 14, 15]:
    letter = get_column_letter(col)
    ws.cell(row=totals_row, column=col, value=f"=SUM({letter}2:{letter}{shopify_end_row})")
    ws.cell(row=totals_row, column=col).font = bf
    ws.cell(row=totals_row, column=col).number_format = nf

if bank_extra:
    totals_row += 1
    ws.cell(row=totals_row, column=8, value="TOTALES BANCO (extra)").font = bf
    for col in [9, 11, 12, 14, 15]:
        letter = get_column_letter(col)
        ws.cell(row=totals_row, column=col, value=f"=SUM({letter}{shopify_end_row + 2}:{letter}{row_idx - 1})")
        ws.cell(row=totals_row, column=col).font = bf
        ws.cell(row=totals_row, column=col).number_format = nf

    totals_row += 1
    ws.cell(row=totals_row, column=8, value="GRAN TOTAL").font = Font(name="Calibri", bold=True, size=11)
    for col in [9, 11, 12, 14, 15]:
        letter = get_column_letter(col)
        ws.cell(row=totals_row, column=col, value=f"=SUM({letter}2:{letter}{row_idx - 1})")
        ws.cell(row=totals_row, column=col).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=totals_row, column=col).number_format = nf

# Column widths
widths = [12, 12, 25, 28, 15, 25, 8, 15, 12, 8, 12, 12, 22, 12, 12, 14, 20]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:Q{row_idx - 1}"

# Save
out = os.path.join(parent, "Negocio", "PEDIDOS-ANTIQUA-2026-COMPLETO.xlsx")
wb.save(out)
print(f"Guardado: {out}")
print(f"Shopify: {shopify_end_row - 1} filas")
print(f"Banco extra: {len(bank_extra)} filas")
print(f"Total: {row_idx - 2} filas")
