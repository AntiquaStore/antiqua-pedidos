"""Generate empty cost catalog Excel for Antiqua."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import models
from catalog import load_catalog

models.init_db()
load_catalog()
products = models.get_all_products()
products.sort(key=lambda p: (p.get("tipo",""), p.get("name","")))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Catalogo de Costes"

# Styles
hf = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
hfill = PatternFill(start_color="212121", end_color="212121", fill_type="solid")
sf = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
sfill = PatternFill(start_color="918A77", end_color="918A77", fill_type="solid")
brd = Border(
    left=Side(style="thin", color="E5E2DC"),
    right=Side(style="thin", color="E5E2DC"),
    top=Side(style="thin", color="E5E2DC"),
    bottom=Side(style="thin", color="E5E2DC"),
)
df = Font(name="Calibri", size=10)
bf = Font(name="Calibri", bold=True, size=10)
nf = '#,##0.00'

# Row 1 - Main headers
merges = [("A1:B1","PIEZA"),("C1:G1","PIEDRAS DE COLOR (LOLA)"),("H1:L1","DIAMANTES"),
          ("M1:O1","ORO"),("P1:S1","TALLER (BARTO)"),("T1:U1","OTROS"),("V1:Y1","TOTALES")]
for rng, txt in merges:
    ws.merge_cells(rng)
    ws[rng.split(":")[0]].value = txt
    ws[rng.split(":")[0]].font = hf
    ws[rng.split(":")[0]].fill = hfill
    ws[rng.split(":")[0]].alignment = Alignment(horizontal="center")
for col in range(1, 26):
    ws.cell(row=1, column=col).fill = hfill
    ws.cell(row=1, column=col).border = brd

# Row 2 - Sub headers
subs = ["Tipo","Nombre",
    "Piedra (tipo)","Piedra (talla)","Piedra (cant.)","Piedra (coste ud.)","Total piedras",
    "Diamante (talla)","Diamante (mm)","Diamante (cant.)","Diamante (coste ud.)","Total diamantes",
    "Peso oro (gr)","Precio oro (EUR/gr)","Total oro",
    "Hechura","Engaste","Rodio/acabado","Total taller",
    "Envio/packaging","Otros",
    "CMV (coste total)","PVP (con IVA)","PVP (sin IVA)","Margen bruto"]
for i, h in enumerate(subs, 1):
    c = ws.cell(row=2, column=i)
    c.value = h
    c.font = sf
    c.fill = sfill
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = brd

# Data rows
for ri, p in enumerate(products, 3):
    ws.cell(row=ri, column=1, value=p.get("tipo","")).font = df
    ws.cell(row=ri, column=2, value=p.get("name","")).font = bf
    # Hints in gray
    pd = p.get("piedras_desc","")
    dd = p.get("diamantes_desc","")
    peso = p.get("peso_gr", 0)
    if pd:
        ws.cell(row=ri, column=3, value=pd).font = Font(name="Calibri", size=9, color="999999")
    if dd:
        ws.cell(row=ri, column=8, value=dd).font = Font(name="Calibri", size=9, color="999999")
    if peso and peso > 0:
        ws.cell(row=ri, column=13, value=peso).number_format = '0.0'
    # Formulas
    ws.cell(row=ri, column=7, value=f"=IF(E{ri}*F{ri},E{ri}*F{ri},0)").number_format = nf
    ws.cell(row=ri, column=12, value=f"=IF(J{ri}*K{ri},J{ri}*K{ri},0)").number_format = nf
    ws.cell(row=ri, column=15, value=f"=M{ri}*N{ri}").number_format = nf
    ws.cell(row=ri, column=19, value=f"=P{ri}+Q{ri}+R{ri}").number_format = nf
    ws.cell(row=ri, column=22, value=f"=G{ri}+L{ri}+O{ri}+S{ri}+T{ri}+U{ri}").number_format = nf
    ws.cell(row=ri, column=22).font = bf
    ws.cell(row=ri, column=24, value=f"=W{ri}/1.21").number_format = nf
    ws.cell(row=ri, column=25, value=f"=X{ri}-V{ri}").number_format = nf
    for col in range(1, 26):
        ws.cell(row=ri, column=col).border = brd

# Widths
widths = {"A":12,"B":22,"C":18,"D":16,"E":10,"F":12,"G":12,"H":16,"I":14,"J":10,"K":12,"L":12,
          "M":10,"N":12,"O":12,"P":12,"Q":12,"R":14,"S":12,"T":14,"U":10,"V":14,"W":14,"X":14,"Y":14}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "C3"

# Sheet 2 - Precio Materiales
ws2 = wb.create_sheet("Precio Materiales")
ws2["A1"] = "TABLA DE PRECIOS DE MATERIALES"
ws2["A1"].font = Font(name="Calibri", bold=True, size=14)

# Diamantes
ws2["A3"] = "DIAMANTES"
ws2["A3"].font = hf
ws2["A3"].fill = hfill
ws2.merge_cells("A3:D3")
for c in ["B3","C3","D3"]:
    ws2[c].fill = hfill
for i, h in enumerate(["Talla","Medida (mm)","Coste unitario (EUR)","Proveedor"], 1):
    c = ws2.cell(row=4, column=i); c.value = h; c.font = sf; c.fill = sfill

diamonds = []
# Brillante: from 1.0 to 6.0 in 0.1mm steps
for i in range(10, 61):
    diamonds.append(("Brillante", f"{i/10:.1f}"))
# Princesa sizes
for s in ["1.5x1.5","2x2","2.5x2.5","3x3","3.5x3.5","3.6x3.6","4x4","5x5"]:
    diamonds.append(("Princesa", s))
# Esmeralda sizes
for s in ["4x3","5x3","5x4","6x4","7x5","8x6"]:
    diamonds.append(("Esmeralda", s))
# Max prices from 2024-2026 invoices
diamond_prices = {
    "1.0": 2.70, "1.1": 4.05, "1.2": 5.80, "1.3": None, "1.4": 8.30,
    "1.5": 8.70, "1.6": 10.90, "1.7": None, "1.8": 16.56, "1.9": None,
    "2.0": 21.50, "2.1": 26.00, "2.2": None, "2.3": 37.90, "2.4": 43.00,
    "2.5": 48.20, "2.6": None, "2.7": None, "2.8": None, "2.9": None,
    "3.0": 93.50, "3.1": None, "3.2": 111.00, "3.3": None, "3.4": None,
    "3.5": 198.05, "3.6": None, "3.7": None, "3.8": None, "3.9": None,
    "4.0": 416.00, "4.3": None,
}
princesa_prices = {"2x2": 42.00}
for i, (t, s) in enumerate(diamonds, 5):
    ws2.cell(row=i, column=1, value=t)
    ws2.cell(row=i, column=2, value=s)
    ws2.cell(row=i, column=4, value="Barto (NOVAO)")
    # Diamantes: leave empty for Barto to fill

# Piedras
ps = 5 + len(diamonds) + 1
ws2.cell(row=ps, column=1, value="PIEDRAS DE COLOR").font = hf
ws2.cell(row=ps, column=1).fill = hfill
ws2.merge_cells(f"A{ps}:E{ps}")
for c in range(2, 6):
    ws2.cell(row=ps, column=c).fill = hfill
for i, h in enumerate(["Piedra","Talla","Medida (mm)","Coste unitario (EUR)","Proveedor"], 1):
    c = ws2.cell(row=ps+1, column=i); c.value = h; c.font = sf; c.fill = sfill

# Stones with max prices from 2024-2026 (price, provider)
stones = [
    ("Zafiro","Carre","1.8x1.8", 5.00, "Lola (Mas Gemas)"),
    ("Zafiro","Carre","2x2", 4.80, "Lola (Mas Gemas)"),
    ("Zafiro","Brillante","1.3", 1.88, "Lola (Mas Gemas)"),
    ("Zafiro","Brillante","3mm", 24.00, "Lola (Mas Gemas)"),
    ("Zafiro","Brillante","4.5mm", 60.00, "Lola (Mas Gemas)"),
    ("Zafiro","Oval","5x4", 60.00, "Lola (Mas Gemas)"),
    ("Zafiro","Oval","6x4", 65.00, "Lola (Mas Gemas)"),
    ("Zafiro","Octogonal","7x5", 300.00, "Lola (Mas Gemas)"),
    ("Esmeralda","Brillante","2.5mm", 12.00, "Lola (Mas Gemas)"),
    ("Esmeralda","Brillante","5mm", 151.20, "Lola (Mas Gemas)"),
    ("Esmeralda","Carre","2x2", 5.20, "Lola (Mas Gemas)"),
    ("Esmeralda","Octogonal","6x4", 150.00, "Lola (Mas Gemas)"),
    ("Esmeralda","Octogonal","7x5", 300.00, "Lola (Mas Gemas)"),
    ("Esmeralda","Octogonal","8x6", 500.00, "Lola (Mas Gemas)"),
    ("Esmeralda","Octogonal","9x7", 500.00, "Lola (Mas Gemas)"),
    ("Esmeralda","Princesa","3x3", 50.00, "Lola (Mas Gemas)"),
    ("Rubi","Brillante","1mm", 1.00, "Lola (Mas Gemas)"),
    ("Rubi","Brillante","2mm", 4.00, "Lola (Mas Gemas)"),
    ("Rubi","Carre","1.8x1.8", 4.00, "Lola (Mas Gemas)"),
    ("Rubi","Carre","3x3", 30.00, "Lola (Mas Gemas)"),
    ("Aguamarina","Princesa","3.5x3.5", 35.00, "Lola (Mas Gemas)"),
    ("Aguamarina","Cojin","4x4", 30.00, "Lola (Mas Gemas)"),
    ("Aguamarina","Pera","6x4", 25.00, "Lola (Mas Gemas)"),
    ("Aguamarina","Canto vivo","7x5", 72.00, "Lola (Mas Gemas)"),
    ("Aguamarina","Oval","6x4", 25.00, "Lola (Mas Gemas)"),
    ("Aguamarina","Baguette","9x7", 60.00, "Lola (Mas Gemas)"),
    ("Turmalina verde","Oval","8x6", 65.00, "Lola (Mas Gemas)"),
    ("Turmalina rosa","Oval","7x5", 38.00, "Lola (Mas Gemas)"),
    ("Granate rodolita","Pera","3x5", 20.00, "Lola (Mas Gemas)"),
    ("Granate rodolita","Octogonal","7x5x3", 20.00, "Lola (Mas Gemas)"),
    ("Granate rodolita","Brillante","1.5mm", 0.20, "Lola (Mas Gemas)"),
    ("Granate rodolita","Brillante","2mm", 2.00, "Lola (Mas Gemas)"),
    ("Amatista","Octogonal","12x7", 30.00, "Lola (Mas Gemas)"),
    ("Amatista verde","Canto vivo","7x5", 5.00, "Lola (Mas Gemas)"),
    ("Amatista verde","Cuadrada","4x4", 1.00, "Lola (Mas Gemas)"),
    ("Citrino","Oval","9x7", 10.00, "Lola (Mas Gemas)"),
    ("Citrino","Oval","12x10", 21.00, "Lola (Mas Gemas)"),
]
for i, (st, cut, sz, price, prov) in enumerate(stones, ps+2):
    ws2.cell(row=i, column=1, value=st)
    ws2.cell(row=i, column=2, value=cut)
    ws2.cell(row=i, column=3, value=sz)
    ws2.cell(row=i, column=4, value=price)
    ws2.cell(row=i, column=4).number_format = '#,##0.00'
    ws2.cell(row=i, column=5, value=prov)

# Oro
os_ = ps + 2 + len(stones) + 1
ws2.cell(row=os_, column=1, value="ORO").font = hf
ws2.cell(row=os_, column=1).fill = hfill
ws2.merge_cells(f"A{os_}:C{os_}")
for c in range(2, 4):
    ws2.cell(row=os_, column=c).fill = hfill
for i, h in enumerate(["Tipo","Precio EUR/gr","Notas"], 1):
    c = ws2.cell(row=os_+1, column=i); c.value = h; c.font = sf; c.fill = sfill
oro_data = [
    ("Oro amarillo 18K", 160.00, "Precio estimado (tendencia alcista). Precio real del dia se captura al recoger pieza"),
    ("Oro blanco 18K", 165.00, "Ligeramente mas caro que amarillo por aleacion"),
    ("Rodio (bano)", 15.00, "Precio por pieza, no por gramo"),
]
for i, (tipo, precio, nota) in enumerate(oro_data, os_+2):
    ws2.cell(row=i, column=1, value=tipo)
    ws2.cell(row=i, column=2, value=precio)
    ws2.cell(row=i, column=2).number_format = '#,##0.00'
    ws2.cell(row=i, column=3, value=nota)

for col in ["A","B","C","D","E"]:
    ws2.column_dimensions[col].width = 20

out = "C:\\Users\\34628\\Desktop\\ANTIQUA\\Negocio\\CATALOGO-COSTES-ANTIQUA-2026.xlsx"
wb.save(out)
print(f"Guardado en: {out}")
print(f"Productos: {len(products)}")
